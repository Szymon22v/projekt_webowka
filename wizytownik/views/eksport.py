"""
eksport.py – Widoki eksportu danych i dynamicznej grafiki (Punkt 4a i 4b).
"""
import csv
import io
import json
from datetime import date, timedelta

from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render

from ..models import Lekarz, Pacjent, Specjalizacja, Wizyta


def strona_eksportu(request):
    return render(request, 'wizytownik/eksport/strona_eksportu.html', {
        'liczba_wizyt':     Wizyta.objects.count(),
        'liczba_lekarzy':   Lekarz.objects.count(),
        'liczba_pacjentow': Pacjent.objects.count(),
    })


# ── Punkt 4a: Eksport CSV ────────────────────────────────────────────────────
def eksport_csv(request):
    typ = request.GET.get('typ', 'wizyty')
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{typ}_export.csv"'
    response.write('\ufeff')  # BOM dla Excela
    writer = csv.writer(response, delimiter=';')

    if typ == 'lekarze':
        writer.writerow(['ID','Tytuł','Imię','Nazwisko','Specjalizacja',
                         'Miasto','Godziny pracy','Cena (zł)','Ocena','Aktywny'])
        for l in Lekarz.objects.select_related('specjalizacja').all():
            writer.writerow([l.pk, l.tytul, l.imie, l.nazwisko, l.specjalizacja.nazwa,
                             l.miasto, l.godz_pracy, l.cena_wizyty, l.ocena,
                             'Tak' if l.aktywny else 'Nie'])
    elif typ == 'pacjenci':
        writer.writerow(['ID','Imię','Nazwisko','E-mail','Telefon','Data rejestracji'])
        for p in Pacjent.objects.all():
            writer.writerow([p.pk, p.imie, p.nazwisko, p.email, p.telefon,
                             p.data_rejestracji.strftime('%d.%m.%Y')])
    else:  # wizyty
        data_od = request.GET.get('data_od', '')
        data_do = request.GET.get('data_do', '')
        qs = Wizyta.objects.select_related('lekarz', 'lekarz__specjalizacja', 'pacjent').all()
        if data_od:
            qs = qs.filter(data__gte=data_od)
        if data_do:
            qs = qs.filter(data__lte=data_do)
        writer.writerow(['ID','Lekarz','Specjalizacja','Pacjent','E-mail pacjenta',
                         'Telefon','Data wizyty','Godzina','Status','Powód wizyty'])
        for w in qs:
            writer.writerow([
                w.pk, w.lekarz.tytul_i_nazwisko, w.lekarz.specjalizacja.nazwa,
                f'{w.pacjent.imie} {w.pacjent.nazwisko}', w.pacjent.email, w.pacjent.telefon,
                w.data.strftime('%d.%m.%Y'), w.godzina.strftime('%H:%M'),
                w.get_status_display(), w.powod_wizyty,
            ])
    return response


# ── Punkt 4a: Eksport XLSX ───────────────────────────────────────────────────
def eksport_xlsx(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('Brak biblioteki openpyxl. Zainstaluj: pip install openpyxl', status=500)

    typ = request.GET.get('typ', 'wizyty')
    wb = openpyxl.Workbook()
    ws = wb.active

    styl_naglowek      = Font(bold=True, color='FFFFFF', size=11)
    wypelnienie_nagl   = PatternFill('solid', fgColor='2563EB')
    wyrownanie_srodek  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ramka              = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    kolory_statusow = {
        'oczekujaca': 'FFF9C4', 'potwierdzona': 'C8E6C9',
        'odbyta': 'E0E0E0',     'anulowana': 'FFCDD2',
    }

    if typ == 'lekarze':
        ws.title = 'Lekarze'
        naglowki = ['ID','Tytuł','Imię','Nazwisko','Specjalizacja','Miasto','Godziny','Cena (zł)','Ocena','Aktywny']
        dane = [[l.pk, l.tytul, l.imie, l.nazwisko, l.specjalizacja.nazwa,
                 l.miasto, l.godz_pracy, float(l.cena_wizyty), float(l.ocena),
                 'Tak' if l.aktywny else 'Nie']
                for l in Lekarz.objects.select_related('specjalizacja').all()]
    else:
        ws.title = 'Wizyty'
        naglowki = ['ID','Lekarz','Specjalizacja','Pacjent','E-mail','Data','Godzina','Status','Powód']
        dane = [[w.pk, w.lekarz.tytul_i_nazwisko, w.lekarz.specjalizacja.nazwa,
                 f'{w.pacjent.imie} {w.pacjent.nazwisko}', w.pacjent.email,
                 w.data.strftime('%d.%m.%Y'), w.godzina.strftime('%H:%M'),
                 w.get_status_display(), w.powod_wizyty]
                for w in Wizyta.objects.select_related('lekarz','lekarz__specjalizacja','pacjent').all()]

    ws.append(naglowki)
    for col_idx in range(1, len(naglowki) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = styl_naglowek
        cell.fill = wypelnienie_nagl
        cell.alignment = wyrownanie_srodek
        cell.border = ramka

    statusy_map = dict(Wizyta._meta.get_field('status').choices)
    for row_idx, wiersz in enumerate(dane, start=2):
        ws.append(wiersz)
        if typ == 'wizyty' and len(wiersz) >= 8:
            kolor = None
            for val, label in statusy_map.items():
                if label == wiersz[7]:
                    kolor = kolory_statusow.get(val)
                    break
            if kolor:
                for col_idx in range(1, len(naglowki) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill('solid', fgColor=kolor)
        for col_idx in range(1, len(naglowki) + 1):
            ws.cell(row=row_idx, column=col_idx).border = ramka

    for col_idx, nagl in enumerate(naglowki, start=1):
        max_len = max(len(str(nagl)), *(len(str(w[col_idx-1])) for w in dane) if dane else [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    bufor = io.BytesIO()
    wb.save(bufor)
    bufor.seek(0)
    response = HttpResponse(
        bufor.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{typ}_export.xlsx"'
    return response


# ── Punkt 4b: Dynamiczna grafika PNG (matplotlib) ───────────────────────────
def wykres_png(request):
    """Generuje wykres jako obraz PNG w locie – bez zapisywania na dysk."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    rodzaj = request.GET.get('rodzaj', 'specjalizacje')
    fig, ax = plt.subplots(figsize=(10, 5))
    fig.patch.set_facecolor('#ffffff')
    ax.set_facecolor('#f8fafc')
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    ax.spines['left'].set_color('#cbd5e1')
    ax.spines['bottom'].set_color('#cbd5e1')

    KOLORY = ['#2563eb','#7c3aed','#db2777','#0891b2','#059669','#d97706','#dc2626','#64748b']

    if rodzaj == 'specjalizacje':
        dane = list(Wizyta.objects.values('lekarz__specjalizacja__nazwa').annotate(ile=Count('id')).order_by('-ile'))
        etykiety = [d['lekarz__specjalizacja__nazwa'] or 'Brak' for d in dane] or ['Brak danych']
        wartosci = [d['ile'] for d in dane] or [0]
        kolory = KOLORY[:len(etykiety)]
        paski = ax.bar(etykiety, wartosci, color=kolory, edgecolor='white', linewidth=0.8, width=0.65)
        for pasek, val in zip(paski, wartosci):
            ax.text(pasek.get_x() + pasek.get_width()/2, pasek.get_height()+0.15,
                    str(val), ha='center', va='bottom', fontsize=9, fontweight='bold', color='#1e293b')
        ax.set_title('Liczba wizyt według specjalizacji', fontsize=13, fontweight='bold', color='#1e293b', pad=15)
        ax.set_ylabel('Liczba wizyt', fontsize=10, color='#64748b')
        ax.yaxis.grid(True, linestyle='--', alpha=0.5, color='#e2e8f0')
        ax.set_axisbelow(True)
        plt.xticks(rotation=20, ha='right', fontsize=9)

    elif rodzaj == 'miesiace':
        dzisiaj = date.today()
        etykiety, wartosci = [], []
        for i in range(5, -1, -1):
            d = dzisiaj.replace(day=1) - timedelta(days=i*30)
            etykiety.append(d.strftime('%b %Y'))
            wartosci.append(Wizyta.objects.filter(data__year=d.year, data__month=d.month).count())
        ax.plot(etykiety, wartosci, color='#2563eb', linewidth=2.5, marker='o', markersize=8,
                markerfacecolor='white', markeredgewidth=2.5, markeredgecolor='#2563eb')
        ax.fill_between(etykiety, wartosci, alpha=0.12, color='#2563eb')
        for x, y in enumerate(wartosci):
            ax.text(x, y+0.25, str(y), ha='center', fontsize=9, fontweight='bold', color='#1e293b')
        ax.set_title('Wizyty w ostatnich 6 miesiącach', fontsize=13, fontweight='bold', color='#1e293b', pad=15)
        ax.set_ylabel('Liczba wizyt', fontsize=10, color='#64748b')
        ax.yaxis.grid(True, linestyle='--', alpha=0.5, color='#e2e8f0')
        ax.set_axisbelow(True)

    elif rodzaj == 'statusy':
        STATUS_KOLORY = {'oczekujaca':'#f59e0b','potwierdzona':'#10b981','odbyta':'#64748b','anulowana':'#ef4444'}
        dane = list(Wizyta.objects.values('status').annotate(ile=Count('id')).order_by('-ile'))
        if not dane:
            dane = [{'status': 'brak', 'ile': 1}]
        statusy_labels_map = dict(Wizyta._meta.get_field('status').choices)
        etykiety = [statusy_labels_map.get(d['status'], d['status']) for d in dane]
        wartosci = [d['ile'] for d in dane]
        kolory   = [STATUS_KOLORY.get(d['status'], '#94a3b8') for d in dane]
        fig.clear()
        ax = fig.add_subplot(111)
        fig.patch.set_facecolor('#ffffff')
        wedges, texts, autotexts = ax.pie(
            wartosci, labels=etykiety, autopct='%1.0f%%', colors=kolory,
            startangle=90, wedgeprops={'edgecolor':'white','linewidth':2},
            pctdistance=0.75, labeldistance=1.12
        )
        for at in autotexts:
            at.set_fontsize(9); at.set_fontweight('bold'); at.set_color('white')
        ax.set_title('Rozkład wizyt według statusu', fontsize=13, fontweight='bold', color='#1e293b', pad=20)

    plt.tight_layout()
    bufor = io.BytesIO()
    plt.savefig(bufor, format='png', dpi=130, bbox_inches='tight')
    plt.close(fig)
    bufor.seek(0)
    return HttpResponse(bufor.read(), content_type='image/png')


def wykres_statystyki(request):
    """Strona ze statystykami: wykresy Chart.js + obraz matplotlib."""
    dane_spec = list(
        Wizyta.objects.values('lekarz__specjalizacja__nazwa').annotate(ile=Count('id')).order_by('-ile')
    )
    top_lekarze = list(
        Lekarz.objects.annotate(ile_wizyt=Count('wizyty')).order_by('-ile_wizyt')[:6]
        .values('imie', 'nazwisko', 'ile_wizyt')
    )
    return render(request, 'wizytownik/eksport/statystyki.html', {
        'spec_labels':   json.dumps([d['lekarz__specjalizacja__nazwa'] for d in dane_spec], ensure_ascii=False),
        'spec_values':   json.dumps([d['ile'] for d in dane_spec]),
        'lekarz_labels': json.dumps([f"{l['imie'][0]}. {l['nazwisko']}" for l in top_lekarze], ensure_ascii=False),
        'lekarz_values': json.dumps([l['ile_wizyt'] for l in top_lekarze]),
    })
